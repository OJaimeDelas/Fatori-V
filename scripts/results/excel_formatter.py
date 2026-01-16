# =============================================================================
# FATORI-V • Results • Excel Formatter
# File: excel_formatter.py
# -----------------------------------------------------------------------------
# Professional formatting utilities for Excel workbooks.
# =============================================================================

from typing import List, Dict, Any
from scripts.logging.logger import log_event

# Check if openpyxl is available
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    log_event('EXCEL_LIBRARY_NOT_AVAILABLE')
    OPENPYXL_AVAILABLE = False


def is_excel_available():
    """
    Check if Excel export is available.
    
    Returns:
        Boolean indicating if openpyxl is installed
    """
    return OPENPYXL_AVAILABLE


def format_header_row(worksheet, row_num=1, freeze=True):
    """
    Format header row with bold font and background color.
    
    Args:
        worksheet: Openpyxl worksheet object
        row_num: Row number to format (default: 1)
        freeze: Whether to freeze panes below this row
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply to all cells in header row
    for cell in worksheet[row_num]:
        if cell.value:  # Only format cells with content
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    # Freeze panes below header
    if freeze:
        freeze_cell = f"A{row_num + 1}"
        worksheet.freeze_panes = freeze_cell


def auto_adjust_columns(worksheet, min_width=8, max_width=50):
    """
    Auto-adjust column widths based on content.
    
    Args:
        worksheet: Openpyxl worksheet object
        min_width: Minimum column width
        max_width: Maximum column width
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        
        for cell in column:
            if column_letter is None:
                column_letter = get_column_letter(cell.column)
            
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        if column_letter:
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_conditional_formatting(worksheet, col_letter, start_row, end_row, rule_type='pass_fail'):
    """
    Apply conditional formatting to a column.
    
    Args:
        worksheet: Openpyxl worksheet object
        col_letter: Column letter (e.g., 'C')
        start_row: Starting row number
        end_row: Ending row number
        rule_type: Type of rule ('pass_fail', 'threshold')
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    from openpyxl.formatting.rule import CellIsRule
    
    if rule_type == 'pass_fail':
        # Green for TRUE/PASS/SUCCESS
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        # Red for FALSE/FAIL
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"
        
        # Rule for success values
        success_rule = CellIsRule(
            operator='equal',
            formula=['"TRUE"'],
            fill=green_fill
        )
        
        # Rule for failure values
        fail_rule = CellIsRule(
            operator='equal',
            formula=['"FALSE"'],
            fill=red_fill
        )
        
        worksheet.conditional_formatting.add(range_str, success_rule)
        worksheet.conditional_formatting.add(range_str, fail_rule)


def format_summary_sheet(worksheet, data):
    """
    Format summary sheet with professional styling.
    
    Args:
        worksheet: Openpyxl worksheet object
        data: Dictionary with summary data
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    # Title styling
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    
    # Find title cell (typically A1)
    if worksheet['A1'].value:
        worksheet['A1'].font = title_font
    
    # Format section headers (cells with content in column A)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if row[0].value and isinstance(row[0].value, str):
            # Check if it looks like a section header (ends with ':')
            if row[0].value.endswith(':'):
                row[0].font = section_font
    
    # Auto-adjust columns
    auto_adjust_columns(worksheet)


def format_session_sheet(worksheet, data):
    """
    Format per-session results sheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        data: List of session dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    # Format header row
    format_header_row(worksheet, row_num=1, freeze=True)
    
    # Apply conditional formatting to success column if present
    if worksheet.max_row > 1:
        # Find success column
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value and 'success' in str(cell.value).lower():
                col_letter = get_column_letter(col_idx)
                apply_conditional_formatting(
                    worksheet,
                    col_letter,
                    start_row=2,
                    end_row=worksheet.max_row,
                    rule_type='pass_fail'
                )
                break
    
    # Auto-adjust columns
    auto_adjust_columns(worksheet)


def add_borders(worksheet, start_row=1, end_row=None, start_col=1, end_col=None):
    """
    Add borders to a range of cells.
    
    Args:
        worksheet: Openpyxl worksheet object
        start_row: Starting row
        end_row: Ending row (None = last row)
        start_col: Starting column
        end_col: Ending column (None = last column)
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    if end_row is None:
        end_row = worksheet.max_row
    
    if end_col is None:
        end_col = worksheet.max_column
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row,
                                    min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border