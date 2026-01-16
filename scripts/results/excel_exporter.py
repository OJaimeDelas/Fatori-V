# =============================================================================
# FATORI-V • Results • Excel Exporter
# File: excel_exporter.py
# -----------------------------------------------------------------------------
# Exports metrics to Excel workbook with multiple formatted sheets.
# =============================================================================

from pathlib import Path
from datetime import datetime
from scripts.results.excel_formatter import (
    is_excel_available,
    format_header_row,
    format_summary_sheet,
    format_session_sheet,
    auto_adjust_columns
)
from scripts.logging import logger

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.log_event('WARNING', warning_message="openpyxl not available - Excel export will be disabled")


def create_summary_sheet(workbook, aggregates, build_metrics):
    """
    Create summary sheet with overall statistics.
    
    Args:
        workbook: Openpyxl workbook object
        aggregates: Aggregated metrics dictionary
        build_metrics: Build metrics dictionary
    
    Returns:
        Worksheet object
    """
    ws = workbook.active
    ws.title = "Summary"
    
    row = 1
    
    # Title
    ws.cell(row, 1, "FATORI-V Run Summary")
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2
    
    # Timestamp
    ws.cell(row, 1, "Generated:")
    ws.cell(row, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    row += 2
    
    # Session statistics
    ws.cell(row, 1, "Execution Summary:")
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    
    ws.cell(row, 1, "Total Sessions")
    ws.cell(row, 2, aggregates.get('session_count', 0))
    row += 1
    
    success = aggregates.get('success', {})
    ws.cell(row, 1, "Successful")
    ws.cell(row, 2, success.get('successful', 0))
    row += 1
    
    ws.cell(row, 1, "Failed")
    ws.cell(row, 2, success.get('failed', 0))
    row += 1
    
    ws.cell(row, 1, "Timeouts")
    ws.cell(row, 2, success.get('timeouts', 0))
    row += 1
    
    ws.cell(row, 1, "Success Rate (%)")
    ws.cell(row, 2, f"{success.get('success_rate_percent', 0):.2f}")
    row += 2
    
    # Duration statistics
    avg_duration = aggregates.get('average_duration_s')
    if avg_duration:
        ws.cell(row, 1, "Average Duration (s)")
        ws.cell(row, 2, f"{avg_duration:.2f}")
        row += 1
    
    total_duration = aggregates.get('total_duration_s')
    if total_duration:
        ws.cell(row, 1, "Total Duration (s)")
        ws.cell(row, 2, f"{total_duration:.2f}")
        row += 2
    
    # Build metrics
    if build_metrics and build_metrics.get('reports_available'):
        ws.cell(row, 1, "Build Metrics:")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        
        timing = build_metrics.get('timing', {})
        if timing:
            ws.cell(row, 1, "WNS (ns)")
            ws.cell(row, 2, timing.get('wns', 'N/A'))
            row += 1
            
            ws.cell(row, 1, "Timing Met")
            ws.cell(row, 2, "TRUE" if timing.get('timing_met', False) else "FALSE")
            row += 1
        
        util = build_metrics.get('utilization', {})
        if util:
            ws.cell(row, 1, "LUTs")
            ws.cell(row, 2, util.get('lut_count', 'N/A'))
            row += 1
            
            ws.cell(row, 1, "FFs")
            ws.cell(row, 2, util.get('ff_count', 'N/A'))
            row += 1
    
    # Format sheet
    format_summary_sheet(ws, {})
    
    return ws


def create_sessions_sheet(workbook, session_metrics):
    """
    Create per-session results sheet.
    
    Args:
        workbook: Openpyxl workbook object
        session_metrics: List of session metrics dictionaries
    
    Returns:
        Worksheet object
    """
    ws = workbook.create_sheet("Per-Session Results")
    
    if not session_metrics:
        ws.cell(1, 1, "No session data available")
        return ws
    
    # Define columns
    columns = [
        ('session_id', 'Session ID'),
        ('benchmark_name', 'Benchmark'),
        ('duration_s', 'Duration (s)'),
        ('status', 'Status'),
        ('success', 'Success'),
        ('timeout_occurred', 'Timeout'),
        ('injection_enabled', 'FI Enabled'),
        ('fi_injection_count', 'Injections'),
        ('benchmark_score', 'Score'),
    ]
    
    # Write headers
    for col_idx, (key, header) in enumerate(columns, 1):
        ws.cell(1, col_idx, header)
    
    # Write data
    for row_idx, metrics in enumerate(session_metrics, 2):
        for col_idx, (key, _) in enumerate(columns, 1):
            value = metrics.get(key, '')
            
            # Format boolean values
            if isinstance(value, bool):
                value = "TRUE" if value else "FALSE"
            
            ws.cell(row_idx, col_idx, value)
    
    # Format sheet
    format_session_sheet(ws, session_metrics)
    
    return ws


def create_build_sheet(workbook, build_metrics):
    """
    Create build metrics sheet.
    
    Args:
        workbook: Openpyxl workbook object
        build_metrics: Build metrics dictionary
    
    Returns:
        Worksheet object
    """
    ws = workbook.create_sheet("Build Metrics")
    
    row = 1
    
    # Headers
    ws.cell(row, 1, "Metric")
    ws.cell(row, 2, "Value")
    format_header_row(ws, row_num=1)
    row += 1
    
    if not build_metrics or not build_metrics.get('reports_available'):
        ws.cell(row, 1, "No build metrics available")
        return ws
    
    # Timing metrics
    timing = build_metrics.get('timing', {})
    if timing:
        ws.cell(row, 1, "Timing")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        
        ws.cell(row, 1, "  WNS (ns)")
        ws.cell(row, 2, timing.get('wns', 'N/A'))
        row += 1
        
        ws.cell(row, 1, "  TNS (ns)")
        ws.cell(row, 2, timing.get('tns', 'N/A'))
        row += 1
        
        ws.cell(row, 1, "  Timing Met")
        ws.cell(row, 2, "TRUE" if timing.get('timing_met', False) else "FALSE")
        row += 1
    
    # Utilization metrics
    util = build_metrics.get('utilization', {})
    if util:
        ws.cell(row, 1, "Utilization")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        
        ws.cell(row, 1, "  LUTs")
        ws.cell(row, 2, util.get('lut_count', 'N/A'))
        row += 1
        
        ws.cell(row, 1, "  FFs")
        ws.cell(row, 2, util.get('ff_count', 'N/A'))
        row += 1
        
        ws.cell(row, 1, "  BRAMs")
        ws.cell(row, 2, util.get('bram_count', 'N/A'))
        row += 1
        
        ws.cell(row, 1, "  DSPs")
        ws.cell(row, 2, util.get('dsp_count', 'N/A'))
        row += 1
    
    # Build duration
    duration = build_metrics.get('build_duration_s')
    if duration:
        ws.cell(row, 1, "Build Duration (s)")
        ws.cell(row, 2, f"{duration:.2f}")
        row += 1
    
    # Auto-adjust columns
    auto_adjust_columns(ws)
    
    return ws


def create_fi_analysis_sheet(workbook, aggregates):
    """
    Create FI analysis sheet (if FI was enabled).
    
    Args:
        workbook: Openpyxl workbook object
        aggregates: Aggregated metrics dictionary
    
    Returns:
        Worksheet object or None if FI not enabled
    """
    fi_count = aggregates.get('fi_enabled_count', 0)
    
    if fi_count == 0:
        return None
    
    ws = workbook.create_sheet("FI Analysis")
    
    row = 1
    
    # Title
    ws.cell(row, 1, "Fault Injection Analysis")
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 2
    
    # Basic statistics
    ws.cell(row, 1, "Sessions with FI")
    ws.cell(row, 2, fi_count)
    row += 1
    
    # Detection metrics
    fi_detection = aggregates.get('fi_detection', {})
    if fi_detection:
        row += 1
        ws.cell(row, 1, "Error Detection:")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        
        ws.cell(row, 1, "Total Injections")
        ws.cell(row, 2, fi_detection.get('total_injections', 0))
        row += 1
        
        ws.cell(row, 1, "Detected Errors")
        ws.cell(row, 2, fi_detection.get('detected_errors', 0))
        row += 1
        
        ws.cell(row, 1, "Undetected Errors")
        ws.cell(row, 2, fi_detection.get('undetected_errors', 0))
        row += 1
        
        ws.cell(row, 1, "Detection Rate (%)")
        ws.cell(row, 2, f"{fi_detection.get('detection_rate_percent', 0):.2f}")
        row += 1
    
    # Coverage metrics
    fi_coverage = aggregates.get('fi_coverage', {})
    if fi_coverage:
        row += 1
        ws.cell(row, 1, "Coverage:")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1
        
        ws.cell(row, 1, "Unique Locations")
        ws.cell(row, 2, fi_coverage.get('unique_location_estimate', 0))
        row += 1
    
    # Auto-adjust columns
    auto_adjust_columns(ws)
    
    return ws


def export_to_excel(metrics_aggregator, output_path, config=None):
    """
    Export metrics to Excel workbook with multiple sheets.
    
    Creates Excel workbook with:
    - Summary sheet
    - Per-session results
    - Build metrics
    - FI analysis (if enabled)
    
    Args:
        metrics_aggregator: MetricsAggregator instance with collected metrics
        output_path: Path where Excel file should be written
        config: Optional configuration dictionary (for additional context)
    
    Returns:
        Boolean indicating success
    """
    if not is_excel_available():
        logger.log_event('ERROR', error_message="openpyxl not installed - cannot export to Excel")
        logger.log_event('ERROR', error_message="Install with: pip install openpyxl")
        return False
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.log_event('DEBUG', debug_message=f"Exporting metrics to Excel: {output_path}")
    
    try:
        # Compute aggregates
        aggregates = metrics_aggregator.compute_aggregates()
        
        # Create workbook
        wb = Workbook()
        
        # Create sheets
        create_summary_sheet(wb, aggregates, metrics_aggregator.build_metrics)
        create_sessions_sheet(wb, metrics_aggregator.session_metrics)
        create_build_sheet(wb, metrics_aggregator.build_metrics)
        create_fi_analysis_sheet(wb, aggregates)
        
        # Save workbook
        wb.save(output_path)
        
        logger.log_event('DEBUG', debug_message=f"Excel workbook created successfully: {output_path}")
        return True
    
    except Exception as e:
        logger.log_event('ERROR', error_message=f"Error creating Excel workbook: {e}")
        return False